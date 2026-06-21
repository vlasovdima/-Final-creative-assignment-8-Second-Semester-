#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Объединяет нагрузку двух семестров в один файл.

Использование:
    python itz_consolidated.py
    python itz_consolidated.py <осень.xls> <весна.xls>
    python itz_consolidated.py <осень.xls> <весна.xls> <выход.xlsx>

По умолчанию:
    осень = Список 15вар - осень.xls
    весна = Список 15вар - весна.xls
    выход = consolidated_output.xlsx
"""

import sys
import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
from pathlib import Path

# ── настройки по умолчанию ──────────────────────────────────
DEFAULT_AUTUMN = "Список 15вар - осень.xls"
DEFAULT_SPRING = "Список 15вар - весна.xls"
DEFAULT_OUT = "consolidated_output.xlsx"
HOURS_PER_RATE = 830.0


# ── читаем один .xls файл ──────────────────────────────────
def read_semester(filepath, semester_name):
    """
    Читает .xls файл семестра (11 колонок A-K):
      A=Дисциплина, B=Вид нагрузки, C=Группы, D=Часы,
      E=Поток, F=Часы по УП, G=Табельный №, H=ФИО,
      I=Должность, J=Аудитория, K=Описание

    Возвращает список словарей. Пустые строки пропускает.
    """
    wb = xlrd.open_workbook(str(filepath))
    ws = wb.sheet_by_index(0)
    rows = []

    for r in range(1, ws.nrows):
        disc = ws.cell_value(r, 0)
        load_type = ws.cell_value(r, 1)
        load = ws.cell_value(r, 3)

        if not disc or not load_type or not load:
            continue

        disc = str(disc).strip()
        load_type = str(load_type).strip()
        if not disc or not load_type:
            continue

        rows.append({
            'semester': semester_name,
            'discipline': disc,
            'load_type': load_type,
            'group': str(ws.cell_value(r, 2)).strip(),
            'load': float(load),
            'flow': str(ws.cell_value(r, 4)).strip(),
            'hours_up': float(ws.cell_value(r, 5) or 0),
            'tab_num': str(ws.cell_value(r, 6)).strip(),
            'fio': str(ws.cell_value(r, 7)).strip(),
            'position': str(ws.cell_value(r, 8)).strip(),
            'audience': str(ws.cell_value(r, 9)).strip(),
            'description': str(ws.cell_value(r, 10)).strip(),
        })

    return rows


# ── считаем часы по преподавателям ─────────────────────────
def calc_teacher_load(all_rows):
    """Суммирует часы по каждому преподавателю. {ФИО: {осень, весна, total}}"""
    data = defaultdict(lambda: {'осень': 0.0, 'весна': 0.0, 'total': 0.0})
    for row in all_rows:
        fio = row['fio']
        if not fio:
            continue
        data[fio][row['semester']] += row['load']
        data[fio]['total'] += row['load']
    return dict(data)


# ── создаём выходной .xlsx ─────────────────────────────────
def write_output(all_rows, teacher_load, filepath):
    """
    Создаёт .xlsx с тремя листами:
      'сводная'       — все строки нагрузки
      'преподаватели' — часы по каждому преподавателю
      'статистика'    — общая статистика
    """
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # ── лист 1: сводная таблица ─────────────────────────────
    ws1 = wb.active
    ws1.title = 'сводная'

    header = ['семестр', 'Дисциплина', 'Вид нагрузки', 'Группы', 'Нагрузка',
              'Поток', 'Количества часов по УП', 'Табельный №', 'ФИО',
              'Должность', 'Желаемая аудитория', 'Описание']
    ws1.append(header)
    for cell in ws1[1]:
        cell.font = bold
        cell.fill = gray
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row in all_rows:
        ws1.append([
            row['semester'], row['discipline'], row['load_type'], row['group'],
            row['load'], row['flow'], row['hours_up'], row['tab_num'],
            row['fio'], row['position'], row['audience'], row['description']
        ])

    for col, w in {'A': 10, 'B': 50, 'C': 25, 'D': 35, 'E': 10, 'F': 10,
                   'G': 12, 'H': 15, 'I': 40, 'J': 20, 'K': 30, 'L': 50}.items():
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = 'A2'

    # ── лист 2: преподаватели ───────────────────────────────
    ws2 = wb.create_sheet('преподаватели')
    ws2.append(['ФИО', 'Осень', 'Весна', 'Всего', 'Отношение к ставке'])
    for cell in ws2[1]:
        cell.font = bold
        cell.fill = gray

    for fio, d in sorted(teacher_load.items(), key=lambda x: -x[1]['total']):
        ws2.append([fio, round(d['осень'], 1), round(d['весна'], 1),
                     round(d['total'], 1), round(d['total'] / HOURS_PER_RATE, 2)])

    ws2.column_dimensions['A'].width = 45
    for col in 'BCDE':
        ws2.column_dimensions[col].width = 12
    ws2.freeze_panes = 'A2'

    # ── лист 3: статистика ──────────────────────────────────
    ws3 = wb.create_sheet('статистика')

    lt_hours = defaultdict(float)
    for row in all_rows:
        lt_hours[row['load_type']] += row['load']

    stats = [
        ['Показатель', 'Значение'],
        ['Всего строк нагрузки', len(all_rows)],
        ['Осень (строк)', sum(1 for r in all_rows if r['semester'] == 'осень')],
        ['Весна (строк)', sum(1 for r in all_rows if r['semester'] == 'весна')],
        ['Всего часов', round(sum(r['load'] for r in all_rows), 1)],
        ['Часы осень', round(sum(r['load'] for r in all_rows if r['semester'] == 'осень'), 1)],
        ['Часы весна', round(sum(r['load'] for r in all_rows if r['semester'] == 'весна'), 1)],
        ['Всего преподавателей', len(teacher_load)],
        ['Норма на 1 ставку', HOURS_PER_RATE],
        [''], ['Тип нагрузки', 'Часы'],
    ]
    for lt, hours in sorted(lt_hours.items(), key=lambda x: -x[1]):
        stats.append([lt, round(hours, 1)])

    for row in stats:
        ws3.append(row)
    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 15

    wb.save(str(filepath))
    print(f"Файл сохранён: {filepath}")


# ── главная функция ────────────────────────────────────────
def main():
    # аргументы командной строки
    if len(sys.argv) == 1:
        path_autumn = Path(DEFAULT_AUTUMN)
        path_spring = Path(DEFAULT_SPRING)
        path_out = Path(DEFAULT_OUT)
    elif len(sys.argv) == 3:
        path_autumn = Path(sys.argv[1])
        path_spring = Path(sys.argv[2])
        path_out = Path(DEFAULT_OUT)
    elif len(sys.argv) >= 4:
        path_autumn = Path(sys.argv[1])
        path_spring = Path(sys.argv[2])
        path_out = Path(sys.argv[3])
    else:
        print("Использование:")
        print("  python itz_consolidated.py")
        print("  python itz_consolidated.py <осень.xls> <весна.xls>")
        print("  python itz_consolidated.py <осень.xls> <весна.xls> <выход.xlsx>")
        return

    if not path_autumn.exists():
        print(f"ОШИБКА: не найден {path_autumn}")
        return
    if not path_spring.exists():
        print(f"ОШИБКА: не найден {path_spring}")
        return

    print("=" * 60)
    print("ФОРМИРОВАНИЕ СВОДНОГО ФАЙЛА НАГРУЗКИ")
    print("=" * 60)

    print(f"\n[1/4] Читаю {path_autumn.name}...")
    autumn = read_semester(path_autumn, 'осень')
    print(f"      {len(autumn)} строк")

    print(f"[2/4] Читаю {path_spring.name}...")
    spring = read_semester(path_spring, 'весна')
    print(f"      {len(spring)} строк")

    all_rows = autumn + spring
    print(f"\n[3/4] Всего строк: {len(all_rows)}")

    missing = [r for r in all_rows if not r['fio']]
    if missing:
        print(f"      ⚠ {len(missing)} строк без ФИО!")
    else:
        print("      ✓ Все ФИО заполнены")

    print("\n[4/4] Считаю нагрузку...")
    load = calc_teacher_load(all_rows)
    print(f"      {len(load)} преподавателей")

    write_output(all_rows, load, path_out)

    print(f"\n{'=' * 60}")
    print(f"{'ФИО':<45} {'Осень':>7} {'Весна':>7} {'Всего':>7} {'Ставка':>7}")
    print("-" * 75)
    for fio, d in sorted(load.items(), key=lambda x: -x[1]['total']):
        print(f"{fio:<45} {d['осень']:>7.1f} {d['весна']:>7.1f} {d['total']:>7.1f} {d['total']/HOURS_PER_RATE:>7.2f}")
    print(f"\nГОТОВО! → {path_out}")
    print("=" * 60)


if __name__ == '__main__':
    main()
